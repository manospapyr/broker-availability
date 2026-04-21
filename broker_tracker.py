#!/usr/bin/env python3
"""
Broker Availability Tracker v3 — Async + Self-Discovering IDs
==============================================================
Architecture:
  • CITY_IDS: static IDs known at write-time (Athens fully populated).
  • city_ids_cache.json: auto-populated on first Playwright run.
  • asyncio.gather with 6 concurrent pages, one browser context.
  • WAF detection → N/A + diagnostic (never ✖ for bot-blocked pages).
  • Brand regex: word-boundary, handles "oto q", "oto-q" etc.
  • Full diagnostics: URL, status, content-len, body head, WAF flag.
"""

import asyncio, base64, json, os, re, sys, time, traceback
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import gspread
import requests
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from playwright.async_api import async_playwright, Page, BrowserContext
    HAS_PW = True
except ImportError:
    HAS_PW = False
    print("WARN: playwright not installed")

# ─── DATES ───────────────────────────────────────────────────────────
def _next_weekday(offset=14):
    d = datetime.now() + timedelta(days=offset)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

PU_D = _next_weekday(14)
DO_D = PU_D + timedelta(days=7)
PU   = PU_D.strftime("%Y-%m-%d")
DO   = DO_D.strftime("%Y-%m-%d")
PUT, DOT, AGE = "10:00", "10:00", 30
print(f"[CONFIG] {PU} → {DO}")

# ─── BRAND REGEX + WAF ───────────────────────────────────────────────
BRAND_RE = {
    "otoQ":    re.compile(r"\boto[\s\-]?q\b", re.I),
    "Drive365": re.compile(r"\bdrive[\s\-]?365\b", re.I),
}
WAF_SIGNALS = ["just a moment","cf-challenge","cf_chl_opt","captcha",
               "are you human","ddos-guard","access denied","403 forbidden","blocked"]

def brand_found(html, brand):   return bool(BRAND_RE[brand].search(html))
def is_waf(html):
    if len(html) < 2000: return True
    lo = html.lower()
    return any(s in lo for s in WAF_SIGNALS)

# ─── CITY IDS ────────────────────────────────────────────────────────
# Athens fully populated from supplied URLs.
# All other cities: iata only. Numeric IDs auto-discovered on first run
# via Playwright form-navigation and cached to city_ids_cache.json.
CITY_IDS = {
    "Athens": {
        "iata": "ATH",
        "enjoytravel": 437,
        "aurum_id": 9514, "aurum_name": "Athen Flughafen",
        "rentcars": 3782,
        "ebookings_plc": 1519, "ebookings_cr": 84,
        "rentcarla": 428,
        "vipcars_country": 62, "vipcars_city": 1259, "vipcars_loc": 482,
        "hotelbeds_piata": 221,
        "hotelbeds_label": "Athens Elefherios Venizelos International Airport (ATH)",
        "discovercars": 1843,
        "yolcu_pid": "ChIJYVzn2RqQoRQRqrPuCt8Vsjg",
        "yolcu_pp":  "37.9415988,23.9477271",
    },
    "Zante":     {"iata":"ZTH"},
    "Chania":    {"iata":"CHQ"},
    "Heraklion": {"iata":"HER"},
    "Valletta":  {"iata":"MLA"},
    "Tirana":    {"iata":"TIA"},
    "Tunis":     {"iata":"TUN"},
    "Enfidha":   {"iata":"NBE"},
    "Monastir":  {"iata":"MIR"},
    "Djerba":    {"iata":"DJE"},
    "Orlando":   {"iata":"MCO"},
    "Miami":     {"iata":"MIA"},
    "Tampa":     {"iata":"TPA"},
    "Hollywood": {"iata":"FLL"},
    "Rabat":     {"iata":"RBA"},
    "Fez":       {"iata":"FEZ"},
    "Tangier":   {"iata":"TNG"},
    "Agadir":    {"iata":"AGA"},
    "Marrakesh": {"iata":"RAK"},
    "Casablanca":{"iata":"CMN"},
    "Podgorica": {"iata":"TGD"},
    "Tivat":     {"iata":"TIV"},
    "Timisoara": {"iata":"TSR"},
    "Plaisance": {"iata":"MRU"},
}

CACHE_PATH = Path("city_ids_cache.json")
_CACHE: dict = {}

def load_cache():
    global _CACHE
    if CACHE_PATH.exists():
        try: _CACHE = json.loads(CACHE_PATH.read_text())
        except Exception: _CACHE = {}

def save_cache():
    CACHE_PATH.write_text(json.dumps(_CACHE, indent=2))

def cids(city):
    base = dict(CITY_IDS.get(city, {}))
    base.update(_CACHE.get(city, {}))
    return base

def cache_set(city, key, val):
    _CACHE.setdefault(city, {})[key] = val
    save_cache()

# ─── AREAS & BROKERS ─────────────────────────────────────────────────
OTOQ_AREAS = {
    "Greece":        ["Athens","Zante","Chania","Heraklion"],
    "Malta":         ["Valletta"],
    "Albania":       ["Tirana"],
    "Tunisia":       ["Tunis","Enfidha","Monastir","Djerba"],
    "United States": ["Orlando","Miami","Tampa","Hollywood"],
    "Morocco":       ["Rabat","Fez","Tangier","Agadir","Marrakesh","Casablanca"],
    "Montenegro":    ["Podgorica"],
    "Romania":       ["Timisoara"],
    "Mauritius":     ["Plaisance"],
}
DRIVE365_AREAS = {
    "Greece":        ["Heraklion","Athens"],
    "Albania":       ["Tirana"],
    "United States": ["Miami","Tampa","Hollywood","Orlando"],
    "Malta":         ["Valletta"],
    "Montenegro":    ["Podgorica","Tivat"],
}
OTOQ_BROKERS = [
    "Discovercars.com","Qeeq.com","Orbitcarhire.com",
    "carrental.hotelbeds.com","Enjoytravel.com","Aurumcars.de",
    "Carjet.com","Rentcars.com","CarFlexi.com",
    "Economybookings.com","Priceline.com","Rentcarla.com",
    "Vipcars.com","Yolcu360","Wisecars.com","BSP-auto.com",
    "StressFreeCarRental.com","otoQ.rent",
]
DRIVE365_BROKERS = [
    "Discovercars.com","Orbitcarhire.com","Vipcars.com",
    "Enjoytravel.com","Carjet.com","Economybookings.com",
    "BSP-auto.com","Aurumcars.de","StressFreeCarRental.com","Drive365.rent",
]

# ─── DIAGNOSTICS ─────────────────────────────────────────────────────
@dataclass
class D:
    broker:str; city:str; brand:str; stage:str; detail:str
    url:str=""; status:int=0; content_len:int=0
    body_head:str=""; waf:bool=False
    ts:str=field(default_factory=lambda:datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

DL: list = []

def dl(broker,city,brand,stage,detail,url="",status=0,content_len=0,body_head="",waf=False):
    DL.append(D(broker=broker,city=city,brand=brand,stage=stage,detail=detail,
                url=url,status=status,content_len=content_len,
                body_head=body_head,waf=waf))

# ─── PLAYWRIGHT HELPERS ──────────────────────────────────────────────
_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")
_STEALTH = """
Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3,4,5]});
Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});
window.chrome={runtime:{}};
"""

async def new_page(ctx):
    p = await ctx.new_page()
    await p.add_init_script(_STEALTH)
    return p

async def pw_get(page, url, wait_ms=3000, broker="", city="", brand=""):
    """Navigate + wait. Returns (html, waf_bool) or (None, False) on error."""
    try:
        r = await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        try:
            await page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        await page.wait_for_timeout(wait_ms)
        html   = await page.content()
        waf    = is_waf(html)
        status = r.status if r else 0
        if waf:
            dl(broker,city,brand,"waf","WAF/bot challenge",url=url,status=status,
               content_len=len(html),body_head=html[:200],waf=True)
        return html, waf
    except Exception as e:
        dl(broker,city,brand,"err",str(e)[:200],url=url)
        return None, False

def result(html, brand, broker, city, url, status=0):
    if html is None or is_waf(html): return "N/A"
    found = brand_found(html, brand)
    dl(broker,city,brand,"ok","✔" if found else "✖",
       url=url,status=status,content_len=len(html),body_head=html[:200])
    return "✔" if found else "✖"

# ─── BROKER AGENTS ───────────────────────────────────────────────────

async def ck_discovercars(page, city, brand):
    ci  = cids(city)
    loc = ci.get("discovercars")
    if not loc:
        dl("Discovercars.com",city,brand,"no_id",f"No discovercars ID for {city}")
        return "N/A"
    sq = base64.b64encode(json.dumps({
        "PickupLocationId":loc,"DropOffLocationId":loc,
        "PickupDateTime":f"{PU}T{PUT}:00","DropOffDateTime":f"{DO}T{DOT}:00",
        "ResidenceCountry":"GR","DriverAge":AGE,"Hash":""
    }).encode()).decode()
    url = f"https://www.discovercars.com/search?sq={sq}"
    html,_ = await pw_get(page,url,4000,"Discovercars.com",city,brand)
    return result(html,brand,"Discovercars.com",city,url)


async def ck_qeeq(page, city, brand):
    """Form-flow only (session hash)."""
    iata = cids(city).get("iata","")
    try:
        await page.goto("https://www.qeeq.com/car/car-rental", wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(2000)
        inp = await page.query_selector("input[id*='pickup'],input[name*='pickup'],input[placeholder*='ick-up']")
        if inp:
            await inp.fill(iata)
            await page.wait_for_timeout(1500)
            sug = await page.query_selector("[class*='suggest'] li,[class*='option'],[role='option']")
            if sug: await sug.click()
        btn = await page.query_selector("button[type='submit'],button[class*='search']")
        if btn: await btn.click()
        await page.wait_for_load_state("networkidle", timeout=15000)
        await page.wait_for_timeout(3000)
        return result(await page.content(),brand,"Qeeq.com",city,page.url)
    except Exception as e:
        dl("Qeeq.com",city,brand,"err",str(e)[:200]); return "N/A"


async def ck_orbit(page, city, brand):
    ci  = cids(city)
    loc = ci.get("vipcars_loc")
    if not loc:
        dl("Orbitcarhire.com",city,brand,"no_id",f"No orbit loc for {city}"); return "N/A"
    url = (f"https://www.orbitcarhire.com/en/reservation/vehicles/"
           f"?pickupDateTime={PU}T{PUT}&dropoffDateTime={DO}T{DOT}"
           f"&pickupLocation={loc}&dropoffLocation={loc}&residenceCountryIso=gr&currency=EUR")
    html,_ = await pw_get(page,url,3000,"Orbitcarhire.com",city,brand)
    return result(html,brand,"Orbitcarhire.com",city,url)


async def ck_hotelbeds(page, city, brand):
    ci    = cids(city)
    piata = ci.get("hotelbeds_piata")
    label = ci.get("hotelbeds_label","")
    if not piata:
        dl("carrental.hotelbeds.com",city,brand,"no_id",f"No hotelbeds PIATA for {city}"); return "N/A"
    url = (f"https://carrental.hotelbeds.com/search.html?VEHICLEGROUP=passenger-vehicle"
           f"&PLABEL={requests.utils.quote(label)}&PIATA={piata}"
           f"&DLABEL={requests.utils.quote(label)}&DIATA={piata}"
           f"&PDAY={PU_D.day}&PMONTHYEAR={PU_D.month}.{PU_D.year}&PDATE={PU}&PTIME={PUT}"
           f"&DDAY={DO_D.day}&DMONTHYEAR={DO_D.month}.{DO_D.year}&DDATE={DO}&DTIME={DOT}"
           f"&AGE={AGE}&CURRENCY=EUR")
    html,_ = await pw_get(page,url,4000,"carrental.hotelbeds.com",city,brand)
    return result(html,brand,"carrental.hotelbeds.com",city,url)


async def _discover_via_form(page, home, iata, url_pattern, cache_keys, city):
    """Generic form-nav + URL extraction helper."""
    try:
        await page.goto(home, wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(2000)
        inp = await page.query_selector(
            "input[name*='pickup'],input[id*='pickup'],input[name*='dep_destination'],"
            "input[placeholder*='ick-up'],input[placeholder*='ick up']")
        if not inp: return None
        await inp.fill(iata)
        await page.wait_for_timeout(2000)
        sug = await page.query_selector(
            "[class*='ui-menu-item'],[class*='suggestion'] li,[role='option'],"
            "[class*='autocomplete'] li")
        if not sug: return None
        await sug.click()
        await page.wait_for_timeout(500)
        btn = await page.query_selector("button[type='submit'],input[type='submit']")
        if btn: await btn.click()
        await page.wait_for_load_state("networkidle", timeout=15000)
        m = re.search(url_pattern, page.url)
        if m:
            vals = [int(g) if g.isdigit() else g for g in m.groups()]
            for key, val in zip(cache_keys, vals):
                cache_set(city, key, val)
            return vals
    except Exception as e:
        print(f"  [discover {home}] {city}: {e}")
    return None


async def ck_enjoytravel(page, city, brand):
    ci   = cids(city)
    ploc = ci.get("enjoytravel")
    if not ploc:
        res = await _discover_via_form(page,
            "https://www.enjoytravel.com/en/car-rental",
            cids(city).get("iata",""),
            r'plocation=(\d+)', ["enjoytravel"], city)
        ploc = res[0] if res else None
    if not ploc:
        dl("Enjoytravel.com",city,brand,"no_id",f"No enjoytravel ID for {city}"); return "N/A"
    url = (f"https://www.enjoytravel.com/en/booking/browse"
           f"?plocation={ploc}&dlocation={ploc}"
           f"&pdate={PU}&ddate={DO}&ptime={PUT}&dtime={DOT}&old=true")
    html,_ = await pw_get(page,url,3000,"Enjoytravel.com",city,brand)
    return result(html,brand,"Enjoytravel.com",city,url)


async def ck_aurum(page, city, brand):
    ci   = cids(city)
    aid  = ci.get("aurum_id")
    name = ci.get("aurum_name","")
    iata = ci.get("iata","")
    if not aid:
        res = await _discover_via_form(page,
            "https://www.aurumcars.de/en/",
            iata, r'dep_destination_id=(\d+)', ["aurum_id"], city)
        if res:
            aid = res[0]
            name = await page.evaluate(
                "()=>{const e=document.querySelector('[name=dep_destination_name]');return e?e.value:''}")
            if name: cache_set(city,"aurum_name",name)
    if not aid:
        dl("Aurumcars.de",city,brand,"no_id",f"No aurum ID for {city}"); return "N/A"
    url = (f"https://www.aurumcars.de/search.php"
           f"?dep_destination_name={requests.utils.quote(name or iata)}"
           f"&dep_destination_id={aid}"
           f"&dest_destination_name=&dest_destination_id="
           f"&dep_date={PU_D.strftime('%d.%m.%Y')}&dep_time={PUT}"
           f"&dest_date={DO_D.strftime('%d.%m.%Y')}&dest_time={DOT}"
           f"&customer_age=27-73")
    html,_ = await pw_get(page,url,3000,"Aurumcars.de",city,brand)
    return result(html,brand,"Aurumcars.de",city,url)


async def ck_carjet(page, city, brand):
    iata = cids(city).get("iata","")
    url  = f"https://www.carjet.com/en/car-hire/{iata.lower()}"
    html,_ = await pw_get(page,url,5000,"Carjet.com",city,brand)
    return result(html,brand,"Carjet.com",city,url)


async def ck_rentcars(page, city, brand):
    ci   = cids(city)
    loc  = ci.get("rentcars")
    iata = ci.get("iata","")
    if not loc:
        res = await _discover_via_form(page,
            "https://www.rentcars.com/en/",
            iata, r'/list/(\d+)-', ["rentcars"], city)
        loc = res[0] if res else None
    if not loc:
        dl("Rentcars.com",city,brand,"no_id",f"No rentcars ID for {city}"); return "N/A"
    pu_ts = int(PU_D.replace(hour=10,minute=0,second=0,microsecond=0).timestamp())
    do_ts = int(DO_D.replace(hour=10,minute=0,second=0,microsecond=0).timestamp())
    url   = f"https://www.rentcars.com/en/booking/list/{loc}-{pu_ts}-{loc}-{do_ts}-0-0-0-0-0-0-0-0"
    html,_ = await pw_get(page,url,4000,"Rentcars.com",city,brand)
    return result(html,brand,"Rentcars.com",city,url)


async def ck_carflexi(page, city, brand):
    """Form-flow only."""
    iata = cids(city).get("iata","")
    try:
        await page.goto("https://www.carflexi.com/en/", wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(2000)
        inp = await page.query_selector(
            "input[name*='pickup'],input[id*='pickup'],input[placeholder*='ick']")
        if inp:
            await inp.fill(iata)
            await page.wait_for_timeout(1500)
            sug = await page.query_selector("[class*='autocomplete'] li,[class*='suggestion'] li")
            if sug: await sug.click()
        for sel,val in [("input[name*='from']",PU),("input[name*='to']",DO)]:
            el = await page.query_selector(sel)
            if el: await el.fill(val)
        btn = await page.query_selector("button[type='submit']")
        if btn: await btn.click()
        await page.wait_for_load_state("networkidle", timeout=15000)
        await page.wait_for_timeout(3000)
        return result(await page.content(),brand,"CarFlexi.com",city,page.url)
    except Exception as e:
        dl("CarFlexi.com",city,brand,"err",str(e)[:200]); return "N/A"


async def ck_ebookings(page, city, brand):
    ci   = cids(city)
    plc  = ci.get("ebookings_plc")
    cr   = ci.get("ebookings_cr")
    iata = ci.get("iata","")
    if not plc or not cr:
        res = await _discover_via_form(page,
            "https://www.economybookings.com/en/",
            iata, r'plc=(\d+).*?cr=(\d+)', ["ebookings_plc","ebookings_cr"], city)
        if res: plc, cr = res[0], res[1]
    if not plc or not cr:
        dl("Economybookings.com",city,brand,"no_id",f"No ebookings IDs for {city}"); return "N/A"
    url = (f"https://www.economybookings.com/en/cars/results"
           f"?cr={cr}&crcy=EUR&lang=en&age={AGE}"
           f"&py={PU_D.year}&pm={PU_D.month:02d}&pd={PU_D.day:02d}"
           f"&dy={DO_D.year}&dm={DO_D.month:02d}&dd={DO_D.day:02d}"
           f"&pt=1000&dt=1000&plc={plc}&dlc={plc}&reload=1")
    html,_ = await pw_get(page,url,4000,"Economybookings.com",city,brand)
    return result(html,brand,"Economybookings.com",city,url)


async def ck_priceline(page, city, brand):
    iata = cids(city).get("iata","")
    url  = (f"https://www.priceline.com/rentalcars/listings/{iata}/{iata}"
            f"/{PU}-{PUT.replace(':','%3A')}/{DO}-{DOT.replace(':','%3A')}/list")
    html,_ = await pw_get(page,url,5000,"Priceline.com",city,brand)
    return result(html,brand,"Priceline.com",city,url)


async def ck_rentcarla(page, city, brand):
    ci   = cids(city)
    loc  = ci.get("rentcarla")
    iata = ci.get("iata","")
    if not loc:
        res = await _discover_via_form(page,
            "https://rentcarla.com/",
            iata, r'/rental-cars/(\d+)/', ["rentcarla"], city)
        loc = res[0] if res else None
    if not loc:
        dl("Rentcarla.com",city,brand,"no_id",f"No rentcarla ID for {city}"); return "N/A"
    pu_ts = int(PU_D.replace(hour=12,minute=0,second=0,microsecond=0).timestamp()*1000)
    pu_s  = PU.replace("-","") + PUT.replace(":","") + "00"
    do_s  = DO.replace("-","") + DOT.replace(":","") + "00"
    url   = f"https://rentcarla.com/rental-cars/{loc}/{loc}/{pu_s}/{do_s}?st={pu_ts}"
    html,_ = await pw_get(page,url,3000,"Rentcarla.com",city,brand)
    return result(html,brand,"Rentcarla.com",city,url)


async def ck_vipcars(page, city, brand):
    ci      = cids(city)
    country = ci.get("vipcars_country")
    vcity   = ci.get("vipcars_city")
    loc     = ci.get("vipcars_loc")
    iata    = ci.get("iata","")
    if not all([country, vcity, loc]):
        res = await _discover_via_form(page,
            "https://www.vipcars.com/",
            iata,
            r'pickup_country=(\d+).*?pickup_city=(\d+).*?pickup_location=(\d+)',
            ["vipcars_country","vipcars_city","vipcars_loc"], city)
        if res: country, vcity, loc = res
    if not all([country, vcity, loc]):
        dl("Vipcars.com",city,brand,"no_id",f"No vipcars IDs for {city}"); return "N/A"
    url = (f"https://www.vipcars.com/search/"
           f"?aff=vipcars_web&language=en&googlemap=1"
           f"&pickup_country={country}&pickup_city={vcity}&pickup_location={loc}"
           f"&dropoff_country={country}&dropoff_city={vcity}&dropoff_location={loc}"
           f"&pickup_date={PU}&pickup_time={PUT}"
           f"&dropoff_date={DO}&dropoff_time={DOT}"
           f"&rc=gr&currency=EUR&drv_age_chk=1&driver_age={AGE}&page=search")
    html,_ = await pw_get(page,url,4000,"Vipcars.com",city,brand)
    return result(html,brand,"Vipcars.com",city,url)


async def ck_yolcu(page, city, brand):
    ci  = cids(city)
    pid = ci.get("yolcu_pid")
    pp  = ci.get("yolcu_pp")
    iata = ci.get("iata","")
    if not pid or not pp:
        res = await _discover_via_form(page,
            "https://yolcu360.com/",
            iata+" airport",
            r'pid=([^&]+).*?p_p=([^&]+)',
            ["yolcu_pid","yolcu_pp"], city)
        if res: pid, pp = res
    if not pid or not pp:
        dl("Yolcu360",city,brand,"no_id",f"No yolcu IDs for {city}"); return "N/A"
    url = (f"https://yolcu360.com/arac-kiralama/search"
           f"?a=30-65&p_d={PU}&d_d={DO}&p_t={PUT}&d_t={DOT}"
           f"&sb=recommended&p_p={pp}&pid={pid}")
    html,_ = await pw_get(page,url,4000,"Yolcu360",city,brand)
    return result(html,brand,"Yolcu360",city,url)


async def ck_wisecars(page, city, brand):
    ci  = cids(city)
    loc = ci.get("vipcars_loc")   # same ID as VipCars
    if not loc:
        dl("Wisecars.com",city,brand,"no_id",f"No wisecars loc for {city}"); return "N/A"
    url = (f"https://www.wisecars.com/en-us/results"
           f"?pick_loc_id={loc}&pick_date={PU}&pick_time={PUT}"
           f"&drop_loc_id={loc}&drop_date={DO}&drop_time={DOT}"
           f"&age={AGE}&rc=US&source=wisecars_web")
    html,_ = await pw_get(page,url,3000,"Wisecars.com",city,brand)
    return result(html,brand,"Wisecars.com",city,url)


async def ck_bsp(page, city, brand):
    """Form-flow only."""
    iata = cids(city).get("iata","")
    try:
        await page.goto("https://www.bsp-auto.com/en/", wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(2000)
        inp = await page.query_selector(
            "input[name*='pick'],input[id*='pick'],input[placeholder*='ick']")
        if inp:
            await inp.fill(iata)
            await page.wait_for_timeout(1500)
            sug = await page.query_selector("[class*='suggestion'] li,[role='option']")
            if sug: await sug.click()
        for sel,val in [("input[name*='from']",PU),("input[name*='to']",DO)]:
            el = await page.query_selector(sel)
            if el: await el.fill(val)
        btn = await page.query_selector("button[type='submit'],input[type='submit']")
        if btn: await btn.click()
        await page.wait_for_load_state("networkidle", timeout=15000)
        await page.wait_for_timeout(3000)
        return result(await page.content(),brand,"BSP-auto.com",city,page.url)
    except Exception as e:
        dl("BSP-auto.com",city,brand,"err",str(e)[:200]); return "N/A"


async def ck_stressfree(page, city, brand):
    """StressFree accepts IATA directly."""
    iata = cids(city).get("iata","")
    url  = (f"https://www.stressfreecarrental.com/en/search-results"
            f"?pickupLocationCode={iata}&dropoffLocationCode={iata}"
            f"&pickupDate={PU_D.strftime('%d%%2F%m%%2F%Y')}"
            f"&dropoffDate={DO_D.strftime('%d%%2F%m%%2F%Y')}"
            f"&pickupHourMinute={PUT}&dropoffHourMinute={DOT}"
            f"&driverStandardAge=on&driverAge={AGE}&language=en&searchType=Airport")
    html,_ = await pw_get(page,url,5000,"StressFreeCarRental.com",city,brand)
    return result(html,brand,"StressFreeCarRental.com",city,url)


async def ck_otoq(page, city, brand):
    iata = cids(city).get("iata","")
    url  = (f"https://www.otoq.rent/en/reservation/vehicles/"
            f"?poslovnica_od={iata}AP%7C"
            f"&date_from={PU_D.strftime('%d%%2F%m%%2F%Y')}&time_from={PUT}"
            f"&date_to={DO_D.strftime('%d%%2F%m%%2F%Y')}&time_to={DOT}#vehicles-list")
    html,_ = await pw_get(page,url,3000,"otoQ.rent",city,brand)
    return result(html,brand,"otoQ.rent",city,url)


async def ck_drive365(page, city, brand):
    iata = cids(city).get("iata","")
    url  = (f"https://www.drive365.rent/en/reservation/vehicles/"
            f"?office_from={iata}AP%7C"
            f"&date_from={PU_D.strftime('%d%%2F%m%%2F%Y')}&time_from={PUT}"
            f"&date_to={DO_D.strftime('%d%%2F%m%%2F%Y')}&time_to={DOT}#vehicles-list")
    html,_ = await pw_get(page,url,3000,"Drive365.rent",city,brand)
    return result(html,brand,"Drive365.rent",city,url)


BROKER_FN = {
    "Discovercars.com":        ck_discovercars,
    "Qeeq.com":                ck_qeeq,
    "Orbitcarhire.com":        ck_orbit,
    "carrental.hotelbeds.com": ck_hotelbeds,
    "Enjoytravel.com":         ck_enjoytravel,
    "Aurumcars.de":            ck_aurum,
    "Carjet.com":              ck_carjet,
    "Rentcars.com":            ck_rentcars,
    "CarFlexi.com":            ck_carflexi,
    "Economybookings.com":     ck_ebookings,
    "Priceline.com":           ck_priceline,
    "Rentcarla.com":           ck_rentcarla,
    "Vipcars.com":             ck_vipcars,
    "Yolcu360":                ck_yolcu,
    "Wisecars.com":            ck_wisecars,
    "BSP-auto.com":            ck_bsp,
    "StressFreeCarRental.com": ck_stressfree,
    "otoQ.rent":               ck_otoq,
    "Drive365.rent":           ck_drive365,
}

# ─── ORCHESTRATOR ────────────────────────────────────────────────────
def build_jobs():
    jobs = []
    seen = set()
    for br in OTOQ_BROKERS:
        for cities in OTOQ_AREAS.values():
            for city in cities:
                k = (br,city,"otoQ")
                if k not in seen: jobs.append(k); seen.add(k)
    for br in DRIVE365_BROKERS:
        for cities in DRIVE365_AREAS.values():
            for city in cities:
                k = (br,city,"Drive365")
                if k not in seen: jobs.append(k); seen.add(k)
    return jobs

async def run_all():
    if not HAS_PW:
        print("Playwright not available"); return {}
    load_cache()
    jobs = build_jobs()
    print(f"\n[RUN] {len(jobs)} checks — 6 concurrent pages")
    results: dict = {}
    res_lock = asyncio.Lock()
    sem = asyncio.Semaphore(6)
    done = [0]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled",
                  "--disable-dev-shm-usage","--disable-gpu","--disable-extensions"])
        ctx = await browser.new_context(
            user_agent=_UA, viewport={"width":1366,"height":768},
            locale="en-US", timezone_id="Europe/Athens")

        async def do_one(br, city, brand):
            async with sem:
                fn = BROKER_FN.get(br)
                if not fn:
                    dl(br,city,brand,"no_agent","No agent"); r = "N/A"
                else:
                    page = await new_page(ctx)
                    try:
                        r = await fn(page, city, brand)
                    except Exception as e:
                        dl(br,city,brand,"exc",str(e)[:200]); r = "N/A"
                    finally:
                        await page.close()
                async with res_lock:
                    done[0] += 1
                    results.setdefault(br,{}).setdefault(city,{})[brand] = r
                    if done[0] % 25 == 0 or done[0] == len(jobs):
                        print(f"  [{done[0]}/{len(jobs)}] {br} / {city} ({brand}) = {r}")

        await asyncio.gather(*[do_one(br,ci,b) for br,ci,b in jobs])
        await browser.close()
    return results

def extract(ar, brand, areas, brokers):
    out = {}
    for br in brokers:
        out[br] = {}
        for cities in areas.values():
            for city in cities:
                out[br][city] = ar.get(br,{}).get(city,{}).get(brand,"N/A")
    return out

# ─── GOOGLE SHEETS ───────────────────────────────────────────────────
def gsc():
    cj = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    if not cj: print("WARN: no GOOGLE_SHEETS_CREDENTIALS"); return None
    return gspread.authorize(Credentials.from_service_account_info(
        json.loads(cj),
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"]))

def sheet_data(label, areas, brokers, res):
    cities = [(co,ci) for co,cl in areas.items() for ci in cl]
    r1 = [label]+[""]*len(cities)
    r2, prev = [""], None
    for co,_ in cities:
        r2.append(co if co!=prev else ""); prev=co
    r3 = [""]+[ci for _,ci in cities]
    rows = [[br]+[res.get(br,{}).get(ci,"N/A") for _,ci in cities] for br in brokers]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    return [r1,r2,r3]+rows+[[],[f"Updated: {ts} | {PU}→{DO}"]]

def diag_data():
    hdr = ["Timestamp","Broker","City","Brand","Stage","Detail",
           "URL","HTTP Status","Content Len","Body Head","WAF"]
    return [hdr]+[[e.ts,e.broker,e.city,e.brand,e.stage,e.detail[:500],
                   e.url,e.status,e.content_len,e.body_head[:200],str(e.waf)]
                  for e in DL]

def update_sheets(oq, d3):
    cl = gsc()
    if not cl: return
    sid = os.environ.get("SPREADSHEET_ID")
    if not sid: print("WARN: no SPREADSHEET_ID"); return
    try:
        sh = cl.open_by_key(sid)
    except Exception as e:
        print(f"ERR open sheet: {e}"); return
    for title, data in [
        ("otoQ",      sheet_data("otoQ",    OTOQ_AREAS,    OTOQ_BROKERS,    oq)),
        ("Drive365",  sheet_data("Drive365",DRIVE365_AREAS,DRIVE365_BROKERS, d3)),
        ("Diagnostics", diag_data()),
    ]:
        try:
            ws = sh.worksheet(title); ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=title,
                                  rows=len(data)+10,
                                  cols=max(len(r) for r in data)+5)
        ws.update(range_name="A1", values=data)
        print(f"  ✔ Sheet '{title}'")

# ─── LOCAL EXCEL ─────────────────────────────────────────────────────
def write_excel(oq, d3, fn):
    wb = Workbook()
    ws1 = wb.active; ws1.title="otoQ"
    _fill_sheet(ws1,"otoQ",OTOQ_AREAS,OTOQ_BROKERS,oq)
    ws2 = wb.create_sheet("Drive365")
    _fill_sheet(ws2,"Drive365",DRIVE365_AREAS,DRIVE365_BROKERS,d3)
    ws3 = wb.create_sheet("Diagnostics"); _fill_diag(ws3)
    wb.save(fn); print(f"  ✔ Excel: {fn}")

def _fill_sheet(ws, label, areas, brokers, res):
    gf=PatternFill("solid",fgColor="C6EFCE"); rf=PatternFill("solid",fgColor="FFC7CE")
    gr=PatternFill("solid",fgColor="D9D9D9"); hd=PatternFill("solid",fgColor="4472C4")
    hw=Font(bold=True,size=11,name="Arial",color="FFFFFF"); df=Font(size=10,name="Arial")
    ct=Alignment(horizontal="center",vertical="center")
    bd=Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    cities=[(co,ci) for co,cl in areas.items() for ci in cl]
    ws.cell(1,1,label).font=Font(bold=True,size=14,name="Arial")
    col=2
    for co,cl in areas.items():
        s=col
        for ci in cl: ws.cell(3,col,ci).alignment=ct; col+=1
        c=ws.cell(2,s,co); c.font=hw; c.fill=hd; c.alignment=ct
        if col-s>1: ws.merge_cells(start_row=2,start_column=s,end_row=2,end_column=col-1)
    for bi,br in enumerate(brokers):
        r=4+bi; ws.cell(r,1,br).font=df
        for ci_i,(_,ci) in enumerate(cities):
            v=res.get(br,{}).get(ci,"N/A"); c=ws.cell(r,ci_i+2,v)
            c.alignment=ct; c.font=df
            c.fill=gf if v=="✔" else rf if v=="✖" else gr
    ws.column_dimensions["A"].width=28
    for x in range(2,col): ws.column_dimensions[get_column_letter(x)].width=14

def _fill_diag(ws):
    for ci,h in enumerate(["Timestamp","Broker","City","Brand","Stage",
                            "Detail","URL","Status","Content Len","Body","WAF"],1):
        ws.cell(1,ci,h).font=Font(bold=True)
    for ri,e in enumerate(DL,2):
        for ci,v in enumerate([e.ts,e.broker,e.city,e.brand,e.stage,e.detail[:500],
                               e.url,e.status,e.content_len,e.body_head[:200],str(e.waf)],1):
            ws.cell(ri,ci,v)

# ─── MAIN ────────────────────────────────────────────────────────────
def main():
    fn = os.environ.get("OUTPUT_FILE","Broker_Availability_Tracker.xlsx")
    print("="*60)
    print("Broker Availability Tracker v3 — Async Self-Discovering")
    print("="*60)
    ar = asyncio.run(run_all())
    oq = extract(ar,"otoQ",   OTOQ_AREAS,   OTOQ_BROKERS)
    d3 = extract(ar,"Drive365",DRIVE365_AREAS,DRIVE365_BROKERS)
    def counts(r):
        v=[x for row in r.values() for x in row.values()]
        return sum(x=="✔" for x in v),sum(x=="✖" for x in v),sum(x=="N/A" for x in v)
    o1,o2,o3=counts(oq); d1,d2,d3_=counts(d3)
    print(f"\n  otoQ:     ✔{o1}  ✖{o2}  N/A {o3}")
    print(f"  Drive365: ✔{d1}  ✖{d2}  N/A {d3_}")
    print(f"  Diag entries: {len(DL)}")
    print("\n--- Excel ---"); write_excel(oq,d3,fn)
    print("--- Sheets ---"); update_sheets(oq,d3)
    print("\nDone!")

if __name__=="__main__":
    main()
